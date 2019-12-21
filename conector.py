'''
Modulo que contiene la conexión para los datos a extraer. En este caso los datos
se extraen desde un excel.

Junto con el conector se definen funciones para transformar los datos a un formato más adecuado

'''

import numpy as np
from openpyxl import load_workbook
wb = load_workbook(filename = 'Proyección LD nuevo 2019.xlsx', data_only=True)

def array2list(array):
    '''
        Devuelve una lista a partir de un array de openpyxl 
    '''

    return [x[0].value for x in array]

def xl2npMatrix(hoja, minrow, maxrow, mincol, maxcol):
    '''
        Toma una hoja de excel y tras definir los limites de extracción de datos,
        (fila minima, fila máxima, columna minima, columna máxima) devuelve una
        matriz de numpy.

        :param hoja: Nombre de la hoja de excel de la que extraen los datos
        :param minrow: Numero de la fila donde se empieza a rescatar los valores.
        :param maxrow: Numero de la fila donde termina de rescatar los valores.
        :param mincol: Numero de la columna donde se empieza a rescatar los valores.
        :param maxcol: Numero de la columna donde termina de rescatar los valores.
        :returns: Matriz de Numpy con los valores.

        >>> xl2npMatrix('hoja1', 1, 3, 2, 4)
        array([[1, 0, 1],
               [1, 0, 1]])
    '''

    listo = list()

    for row in hoja.iter_rows(min_row = minrow, max_row = maxrow, 
    min_col = mincol, max_col = maxcol):
        for cell in row:
            listo.append(cell.value)
    
    listo = np.array(listo)
    listo = listo.reshape(maxrow - minrow + 1, maxcol - mincol + 1)

    return listo